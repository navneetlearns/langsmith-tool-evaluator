#!/usr/bin/env python3
"""
copilot_query_pipeline.py  —  Multi-account Copilot Eval Pipeline

HEART.md principles:
1. Auto-OTP auth (user never sees this — internal)
2. Sends every query via SSE stream, NO RETRY — captures real user experience
3. Tracks response_time_seconds per query
4. Versioned output: query_results_v{N}.jsonl (never overwrites past runs)
5. Supports tool selection accuracy + step count metrics
6. Incremental writes — each query appended immediately (inspectable mid-run)
7. Resume capability — skip already-completed queries

Usage:
    python3 copilot_query_pipeline.py --account surana              # new run (auto-version)
    python3 copilot_query_pipeline.py --account hirafoods --run 3   # explicit version
    python3 copilot_query_pipeline.py --account hirafoods --resume 2  # resume from v2
"""

import json
import os
import sys
import time
import uuid
import urllib.request
import urllib.error
import base64
from pathlib import Path
from datetime import datetime

# Force unbuffered output
sys.stdout.reconfigure(line_buffering=True)
sys.stderr.reconfigure(line_buffering=True)

SCRIPT_DIR = Path(__file__).parent.resolve()
LOG_FILE = SCRIPT_DIR / "pipeline_run.log"


def log_msg(msg: str):
    with open(LOG_FILE, "a") as f:
        f.write(f"{datetime.now().isoformat()} | {msg}\n")
    print(msg, flush=True)


# ============================================================
# CONFIG LOADER  (simple YAML — no external deps needed)
# ============================================================

def load_account_config(account_name: str) -> dict:
    """Load account config from accounts/<name>/config.yaml.
    
    Uses a simple line-by-line parser — avoids PyYAML dependency.
    Supports nested dicts via indentation.
    """
    config_path = SCRIPT_DIR / "accounts" / account_name / "config.yaml"
    if not config_path.exists():
        log_msg(f"Config not found: {config_path}")
        sys.exit(1)

    cfg = {}
    current_section = cfg
    section_stack = []

    with open(config_path) as f:
        for raw_line in f:
            line = raw_line.rstrip()
            if not line or line.startswith("#"):
                continue

            # Detect indent level
            stripped = line.lstrip()
            indent = len(line) - len(stripped)

            if ":" not in stripped:
                continue

            key, _, value = stripped.partition(":")
            key = key.strip()
            value = value.strip().strip('"').strip("'")

            # Strip inline comments (everything after # outside quotes)
            if "#" in value:
                # Only strip if # is not inside a UUID or URL path
                hash_pos = value.find("#")
                if hash_pos > 0 and value[hash_pos-1] == " ":
                    value = value[:hash_pos].strip().strip('"').strip("'")

            if indent == 0:
                current_section = cfg
                section_stack = []
                if value == "":
                    # Start of a nested section
                    current_section[key] = {}
                    current_section = current_section[key]
                    section_stack.append(key)
                else:
                    current_section[key] = value
            else:
                # Nested under the last section
                if value == "":
                    # Further nesting
                    if key not in current_section:
                        current_section[key] = {}
                    current_section = current_section[key]
                    section_stack.append(key)
                else:
                    current_section[key] = value

    # Post-process: convert numeric values
    for k in ("sse_timeout", "sse_read_timeout"):
        if k in cfg:
            cfg[k] = int(cfg[k])

    cfg["account_dir"] = SCRIPT_DIR / "accounts" / account_name
    cfg["runs_dir"] = cfg["account_dir"] / "runs"
    cfg["manifest_file"] = cfg["runs_dir"] / "manifest.json"
    cfg["excel_file"] = cfg["account_dir"] / "queries.xlsx"

    return cfg


# ============================================================
# VERSIONING
# ============================================================

def get_next_version(runs_dir: Path) -> int:
    runs_dir.mkdir(parents=True, exist_ok=True)
    existing = sorted(runs_dir.glob("query_results_v*.jsonl"))
    versions = []
    for f in existing:
        try:
            v = int(f.stem.split("_v")[1])
            versions.append(v)
        except (IndexError, ValueError):
            continue
    return max(versions) + 1 if versions else 1


def load_manifest(manifest_file: Path) -> dict:
    if manifest_file.exists():
        return json.loads(manifest_file.read_text())
    return {"runs": []}


def save_manifest(manifest_file: Path, manifest: dict) -> None:
    manifest_file.write_text(json.dumps(manifest, indent=2, ensure_ascii=False))


def update_manifest(manifest_file: Path, version: int, output_file: Path,
                    total: int, success: int, fail: int, avg_time: float) -> None:
    manifest = load_manifest(manifest_file)
    manifest["runs"].append({
        "version": version,
        "file": output_file.name,
        "timestamp": datetime.now().isoformat(),
        "total_queries": total,
        "success": success,
        "failed": fail,
        "avg_response_time_seconds": round(avg_time, 2),
    })
    save_manifest(manifest_file, manifest)


# ============================================================
# AUTH
# ============================================================

BASE_HEADERS = {
    "accept": "*/*",
    "content-type": "application/json",
    "user-agent": "hermes-pipeline/2.0",
    "origin": "https://copilot.zotok.ai",
    "referer": "https://copilot.zotok.ai/",
}


class CopilotAuth:
    """OTP-based authentication with auto-refresh."""

    def __init__(self, phone: str, base_url: str):
        self.phone = phone
        self.base_url = base_url
        self.token: str | None = None
        self.refresh_token: str | None = None
        self.token_expires_at: float = 0

    def _api_call(self, method: str, path: str, body: dict | None = None,
                  auth: bool = False) -> tuple[int, dict]:
        url = f"{self.base_url}{path}"
        data = json.dumps(body).encode() if body else None
        headers = dict(BASE_HEADERS)
        if auth and self.token:
            headers["authorization"] = f"Bearer {self.token}"
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                content = resp.read().decode()
                return resp.status, json.loads(content)
        except urllib.error.HTTPError as e:
            body_text = e.read().decode()
            try:
                return e.code, json.loads(body_text)
            except Exception:
                return e.code, {"_error": body_text[:300]}
        except Exception as e:
            return 0, {"_error": str(e)}

    def login(self) -> str | None:
        print(f"  [auth] Sending OTP to {self.phone}...")
        status, resp = self._api_call("POST", "/hub/orgs/api/copilot/sendOtp",
                                      {"mobile": self.phone})
        if status not in (200, 201):
            print(f"  [auth] sendOtp failed: {status} {resp.get('_error', resp)}")
            return None

        data = resp.get("data", {})
        otp = data.get("otp")
        otp_token = data.get("otpToken")
        flow = data.get("flow", "?")
        if not otp:
            print(f"  [auth] No OTP received. Response: {resp}")
            return None

        print(f"  [auth] OTP received ({flow} flow), verifying...")
        status, resp = self._api_call("POST", "/hub/orgs/api/copilot/verifyOtp", {
            "mobile": self.phone,
            "otp": otp,
            "otpToken": otp_token,
        })
        if status not in (200, 201):
            print(f"  [auth] verifyOtp failed: {status} {resp.get('_error', resp)}")
            return None

        data = resp.get("data", {})
        token = data.get("token") or data.get("accessToken")
        self.refresh_token = data.get("refreshToken")
        if not token:
            print(f"  [auth] No token in response. Flow: {data.get('flow')}.")
            return None

        self.token = token
        self._decode_expiry(token)
        print(f"  [auth] Authenticated. Token valid for {self._mins_remaining():.0f} min")
        return token

    def refresh(self) -> str | None:
        if not self.refresh_token:
            print("  [auth] No refresh token, re-logging in...")
            return self.login()

        print(f"  [auth] Refreshing token...")
        status, resp = self._api_call("POST", "/hub/orgs/api/copilot/refresh-token",
                                      {"refreshToken": self.refresh_token})
        if status not in (200, 201):
            print(f"  [auth] Refresh failed ({status}), re-logging in...")
            return self.login()

        data = resp.get("data", {})
        new_token = data.get("token") or data.get("accessToken")
        if new_token:
            self.token = new_token
            self._decode_expiry(new_token)
            print(f"  [auth] Token refreshed. Valid for {self._mins_remaining():.0f} min")
        return new_token

    def ensure_token(self) -> str:
        if not self.token:
            if not self.login():
                raise RuntimeError("Authentication failed")
        elif self._mins_remaining() < 2:
            print(f"  [auth] Token expiring ({self._mins_remaining():.0f} min), refreshing...")
            if not self.refresh():
                raise RuntimeError("Token refresh failed")
        return self.token

    def invalidate(self):
        self.token = None

    def _decode_expiry(self, token: str) -> None:
        try:
            parts = token.split(".")
            if len(parts) >= 2:
                payload_b64 = parts[1]
                padding = 4 - len(payload_b64) % 4
                if padding != 4:
                    payload_b64 += "=" * padding
                payload = json.loads(base64.b64decode(payload_b64))
                self.token_expires_at = payload.get("exp", 0)
        except Exception:
            self.token_expires_at = 0

    def _mins_remaining(self) -> float:
        return max(0, (self.token_expires_at - time.time()) / 60)


# ============================================================
# API CLIENT
# ============================================================

class CopilotClient:
    """Client for the ZoTok Copilot API."""

    def __init__(self, auth: CopilotAuth, cfg: dict):
        self.auth = auth
        self.cfg = cfg

    def _request(self, method: str, path: str, body: dict | None = None,
                 timeout: int = 30) -> tuple[int, dict | str]:
        token = self.auth.ensure_token()
        url = f"{self.cfg['base_url']}{path}"
        data = json.dumps(body).encode() if body else None
        headers = dict(BASE_HEADERS)
        headers["authorization"] = f"Bearer {token}"
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                content = resp.read()
                try:
                    return resp.status, json.loads(content.decode())
                except Exception:
                    return resp.status, content.decode()
        except urllib.error.HTTPError as e:
            body_text = e.read().decode()
            try:
                return e.code, json.loads(body_text)
            except Exception:
                return e.code, body_text
        except Exception as e:
            return 0, str(e)

    def init_thread(self, thread_id: str) -> bool:
        status, _ = self._request("POST", "/hub/copilot/threads/init", {
            "thread_id": thread_id,
            "sellerWorkspaceId": self.cfg["workspace_id"],
        })
        return status in (200, 201)

    def stream_query(self, thread_id: str, message: str) -> dict:
        """Send a query via /stream and parse SSE response. NO RETRY."""
        body = {
            "thread_id": thread_id,
            "message": message,
            "sellerWorkspaceId": self.cfg["workspace_id"],
            "wa_config_id": self.cfg["wa_config_id"],
            "seller_details": self.cfg["seller_details"],
            "llm_provider": self.cfg["llm_provider"],
        }

        token = self.auth.ensure_token()
        url = f"{self.cfg['base_url']}/hub/copilot/stream"
        data = json.dumps(body).encode()
        headers = dict(BASE_HEADERS)
        headers["authorization"] = f"Bearer {token}"

        result = {
            "response": "",
            "tool_calls": [],
            "status_sequence": [],
            "suggestions": [],
            "response_time_seconds": 0.0,
            "step_count": 0,
            "error": None,
        }

        query_start = time.time()
        sse_timeout = self.cfg.get("sse_timeout", 300)
        sse_read_timeout = self.cfg.get("sse_read_timeout", 120)

        try:
            req = urllib.request.Request(url, data=data, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=sse_timeout) as resp:
                buffer = ""
                event_type = None
                last_data_time = time.time()

                while True:
                    if time.time() - last_data_time > sse_read_timeout:
                        result["error"] = f"SSE hang: no data for {sse_read_timeout}s"
                        break

                    chunk = resp.read(4096)
                    if not chunk:
                        break

                    last_data_time = time.time()
                    buffer += chunk.decode(errors="replace")

                    while "\n\n" in buffer:
                        event_text, buffer = buffer.split("\n\n", 1)
                        event_type = None
                        event_data = ""

                        for line in event_text.split("\n"):
                            if line.startswith("event: "):
                                event_type = line[7:].strip()
                            elif line.startswith("data: "):
                                event_data += line[6:]

                        if event_type and event_data:
                            result["status_sequence"].append(event_type)

                            try:
                                parsed = json.loads(event_data)
                            except json.JSONDecodeError:
                                parsed = {}

                            if event_type == "tool_start":
                                tool_name = parsed.get("tool") or parsed.get("tool_name") or parsed.get("name", "?")
                                tool_input = parsed.get("input") or parsed.get("arguments") or {}
                                result["tool_calls"].append({"tool": tool_name, "input": tool_input})

                            elif event_type == "message":
                                if not result["response"]:
                                    result["response"] = parsed.get("content", "")

                            elif event_type == "token":
                                # Streaming token — accumulate into response
                                token_text = parsed.get("content") or parsed.get("token") or ""
                                if isinstance(token_text, str):
                                    result["response"] += token_text

                            elif event_type == "ui":
                                # UI rendering event — may contain structured output
                                ui_content = parsed.get("content") or parsed.get("data")
                                if ui_content and isinstance(ui_content, str):
                                    result["response"] += ui_content

                            elif event_type == "suggestions":
                                s = parsed.get("suggestions") or parsed.get("data") or []
                                if isinstance(s, list):
                                    result["suggestions"] = s

        except urllib.error.HTTPError as e:
            body_text = e.read().decode(errors="replace")[:300]
            result["error"] = f"HTTP {e.code}: {body_text}"
            if e.code == 401:
                self.auth.invalidate()

        except Exception as e:
            result["error"] = str(e)

        result["response_time_seconds"] = round(time.time() - query_start, 2)
        result["step_count"] = len(result["status_sequence"])
        return result


# ============================================================
# EXCEL PARSER  (columns: Query, Expected Response, Remarks, Expected Tool)
# ============================================================

def parse_chat_queries(excel_path: str) -> list[dict]:
    """Parse test queries from Excel. Supports two formats:

    Format A (Surana): Sheet "Chat Queries", bold rows = category headers.
        Col A: Query, Col B: Expected Response, Col C: Remarks,
        Col D: Expected Tool (optional)

    Format B (Unifoods): Sheet "Sheet1", column-based categories.
        Col A: No., Col B: Scenario (category), Col C: Query,
        Col D: Expected Tools (optional)
    """
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Detect format
    if "Chat Queries" in wb.sheetnames:
        return _parse_format_a(wb)
    elif "Sheet1" in wb.sheetnames:
        return _parse_format_b(wb)
    else:
        # Fall back to first sheet as Format A
        return _parse_format_a(wb)


def _parse_format_a(wb) -> list[dict]:
    """Surana format: bold rows = category headers."""
    ws = wb["Chat Queries"] if "Chat Queries" in wb.sheetnames else wb[wb.sheetnames[0]]

    queries = []
    current_category = "General"

    for row_idx in range(2, ws.max_row + 1):
        query = ws.cell(row_idx, 1).value
        response = ws.cell(row_idx, 2).value
        remarks = ws.cell(row_idx, 3).value
        expected_tool = ws.cell(row_idx, 4).value

        cell = ws.cell(row_idx, 1)
        is_bold = bool(cell.font and cell.font.bold)

        if is_bold and query:
            current_category = str(query).strip()
            continue
        if not query:
            continue

        q = {
            "query": str(query).strip(),
            "copilot_response": str(response).strip() if response else "",
            "remarks": str(remarks).strip() if remarks else "",
            "category": current_category,
        }
        if expected_tool and str(expected_tool).strip():
            q["expected_tool"] = str(expected_tool).strip()
        queries.append(q)

    wb.close()
    return queries


def _parse_format_b(wb) -> list[dict]:
    """Unifoods format: column B = category, column C = query, column D = expected tools."""
    ws = wb["Sheet1"]

    queries = []
    for row_idx in range(2, ws.max_row + 1):
        scenario = ws.cell(row_idx, 2).value   # Col B: category
        query = ws.cell(row_idx, 3).value      # Col C: query
        expected_tool = ws.cell(row_idx, 4).value  # Col D: expected tools

        if not query:
            continue

        q = {
            "query": str(query).strip(),
            "copilot_response": "",
            "remarks": "",
            "category": str(scenario).strip() if scenario else "General",
        }
        if expected_tool and str(expected_tool).strip():
            q["expected_tool"] = str(expected_tool).strip()
        queries.append(q)

    wb.close()
    return queries


# ============================================================
# MAIN PIPELINE
# ============================================================

def main():
    # Parse CLI flags
    account = "surana"  # default
    run_version = None  # None = auto-increment
    resume = False
    single_query = None
    
    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == "--account" and i + 1 < len(args):
            account = args[i + 1]
            i += 2
        elif args[i].startswith("--account="):
            account = args[i].split("=", 1)[1]
            i += 1
        elif args[i] == "--run" and i + 1 < len(args):
            run_version = int(args[i + 1])
            i += 2
        elif args[i].startswith("--run="):
            run_version = int(args[i].split("=", 1)[1])
            i += 1
        elif args[i] == "--resume" and i + 1 < len(args):
            resume = int(args[i + 1])
            i += 2
        elif args[i].startswith("--resume="):
            resume = int(args[i].split("=", 1)[1])
            i += 1
        elif args[i] == "--query" and i + 1 < len(args):
            single_query = args[i + 1]
            i += 2
        elif args[i].startswith("--query="):
            single_query = args[i].split("=", 1)[1]
            i += 1
        else:
            i += 1

    # Load config
    cfg = load_account_config(account)
    log_msg("=" * 70)
    log_msg(f"  COPILOT QUERY PIPELINE — {cfg['account_name']}")
    log_msg("  Principles: No retry | Patient SSE | Response timing | Versioned")
    log_msg("=" * 70)

    # Determine version and output file
    if resume:
        version = resume
        output_file = cfg["runs_dir"] / f"query_results_v{version}.jsonl"
        if not output_file.exists():
            log_msg(f"  ERROR: Cannot resume v{version}, file not found: {output_file}")
            sys.exit(1)
        log_msg(f"  Account: {account}")
        log_msg(f"  Resuming version: v{version}")
        log_msg(f"  Output: {output_file}")
    else:
        version = run_version if run_version else get_next_version(cfg["runs_dir"])
        output_file = cfg["runs_dir"] / f"query_results_v{version}.jsonl"
        cfg["runs_dir"].mkdir(parents=True, exist_ok=True)
        log_msg(f"  Account: {account}")
        log_msg(f"  Run version: v{version}")
        log_msg(f"  Output: {output_file}")

    # Step 1: Parse Excel queries
    excel_path = cfg["excel_file"]
    log_msg(f"Reading test queries from: {excel_path}")
    if not excel_path.exists():
        log_msg(f"  Excel not found: {excel_path}")
        sys.exit(1)

    queries = parse_chat_queries(str(excel_path))
    log_msg(f"  Found {len(queries)} queries across categories")
    categories = {}
    for q in queries:
        categories[q["category"]] = categories.get(q["category"], 0) + 1
    for cat, count in sorted(categories.items()):
        log_msg(f"    {cat}: {count} queries")

    # Count queries with expected_tool
    expected_count = sum(1 for q in queries if q.get("expected_tool"))
    if expected_count:
        log_msg(f"  {expected_count}/{len(queries)} queries have expected_tool (tool accuracy will be tracked)")

    # Step 2: Authenticate
    log_msg("Authenticating to Copilot API...")
    auth = CopilotAuth(cfg["phone"], cfg["base_url"])
    try:
        token = auth.login()
        if not token:
            log_msg("  Authentication failed!")
            sys.exit(1)
    except Exception as e:
        log_msg(f"  Auth error: {e}")
        sys.exit(1)

    client = CopilotClient(auth, cfg)

    # Step 3: Process each query
    # If resuming, load existing results and skip completed queries
    completed_indices = set()
    results = []
    
    if resume:
        log_msg(f"  Loading existing results from v{resume}...")
        with open(output_file) as f:
            for line in f:
                if line.strip():
                    rec = json.loads(line)
                    results.append(rec)
                    completed_indices.add(rec.get("query_index"))
        log_msg(f"  Found {len(results)} existing results ({len(completed_indices)} queries completed)")
    
    log_msg("Processing queries (no retry, patient SSE, timing each response)...")

    total = len(queries)
    success_count = sum(1 for r in results if not r.get("error"))
    fail_count = sum(1 for r in results if r.get("error"))

    start_time = time.time()

    for idx, q in enumerate(queries, 1):
        query_text = q["query"]
        category = q["category"]
        elapsed = time.time() - start_time
        eta = (elapsed / max(1, idx - 1)) * (total - idx + 1) if idx > 1 else 0

        # Skip if already completed (resume mode)
        if idx in completed_indices:
            log_msg(f"  [{idx}/{total}] ({category[:22]:22s}) {query_text[:55]:.55s} — SKIPPED (already completed)")
            continue

        log_msg(f"  [{idx}/{total}] ({category[:22]:22s}) {query_text[:55]:.55s}")

        # Ensure auth token
        try:
            auth.ensure_token()
        except Exception as e:
            log_msg(f"         [FAIL] Auth refresh failed: {e}")
            results.append(make_fail_record(idx, q, version, f"Auth failed: {e}"))
            fail_count += 1
            continue

        # Init thread
        thread_id = str(uuid.uuid4())
        if not client.init_thread(thread_id):
            log_msg(f"         [FAIL] Thread init failed")
            results.append(make_fail_record(idx, q, version, "Thread init failed", thread_id))
            fail_count += 1
            continue

        # Stream query (NO RETRY)
        result = client.stream_query(thread_id, query_text)

        # Build record
        record = {
            "query_index": idx,
            **q,
            "thread_id": thread_id,
            "tool_calls": result.get("tool_calls", []),
            "response": result.get("response", ""),
            "status_sequence": result.get("status_sequence", []),
            "suggestions": result.get("suggestions", []),
            "response_time_seconds": result.get("response_time_seconds", 0.0),
            "step_count": result.get("step_count", 0),
            "error": result.get("error"),
            "timestamp": datetime.now().isoformat(),
            "run_version": version,
        }
        results.append(record)

        # Log result
        error = record.get("error")
        resp_time = record.get("response_time_seconds", 0)
        # Write this query result immediately (incremental)
        with open(output_file, "a") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
        
        error = record.get("error")
        resp_time = record.get("response_time_seconds", 0)
        
        if error:
            fail_count += 1
            log_msg(f"         [FAIL] {error}  (time: {resp_time}s)")
        else:
            tools = record.get("tool_calls", [])
            resp_len = len(record.get("response", ""))
            tool_names = [t.get("tool", "?") for t in tools]
            steps = record.get("step_count", 0)
            log_msg(f"         [OK]   tools={tool_names} steps={steps} resp_len={resp_len}  (time: {resp_time}s)")
            success_count += 1

        time.sleep(1)

    # Step 4: Update manifest (JSONL already written incrementally)
    log_msg(f"Updating manifest for v{version}...")
    successful_results = [r for r in results if not r.get("error")]
    avg_time = (sum(r["response_time_seconds"] for r in successful_results) / len(successful_results)
                if successful_results else 0.0)

    update_manifest(cfg["manifest_file"], version, output_file,
                    total=len(queries), success=success_count, fail=fail_count,
                    avg_time=avg_time)

    # Step 5: Summary
    elapsed_total = time.time() - start_time
    log_msg("=" * 70)
    log_msg(f"  PIPELINE COMPLETE — v{version}")
    log_msg(f"  Account: {account} ({cfg['account_name']})")
    log_msg(f"  Queries: {len(queries)} | Success: {success_count} | Failed: {fail_count}")
    log_msg(f"  Avg response time: {avg_time:.1f}s")
    log_msg(f"  Total wall time: {elapsed_total/60:.1f} min")
    if resume:
        log_msg(f"  Resumed from v{resume}, completed {len(queries) - len(completed_indices)} new queries")
    log_msg(f"  Output: {output_file}")
    log_msg("=" * 70)


def make_fail_record(idx: int, q: dict, version: int, error: str,
                     thread_id: str | None = None) -> dict:
    return {
        "query_index": idx,
        **q,
        "thread_id": thread_id,
        "tool_calls": [],
        "response": "",
        "status_sequence": [],
        "suggestions": [],
        "response_time_seconds": 0.0,
        "step_count": 0,
        "error": error,
        "timestamp": datetime.now().isoformat(),
        "run_version": version,
    }


if __name__ == "__main__":
    main()
